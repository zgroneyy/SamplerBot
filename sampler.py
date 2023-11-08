#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Sampling Methods and File Operations Script

Author: Ozgur
Date: November 8, 2023

This script provides various sampling methods (random, stratified, systematic, and cluster) for Excel and CSV files.
Additionally, it offers functions for creating empty Excel files and writing data into files.

Usage:
- Create empty Excel files with create_empty_excel_file(filename).
- Print data into files in either XLSX or CSV format with print_into_file(data_array, file_type).
- Check if a file is in the supported format (XLSX or CSV) using check_file_format(filename).
- Create a test Excel file with random data using create_test_file(filename, num_columns, num_rows).
- Perform random sampling from an Excel file using random_sampler(filename, sampling_set_size, sheet_name).
- Perform stratified sampling from an Excel or CSV file using stratified_sampler(filename, groupby_column_num, sample_size).
- Perform systematic sampling from an Excel file using systematic_sampler(sampling_set_size, filename, sheet_name).
- Perform cluster sampling from an Excel file using cluster_sampler(sampling_set_size, sampling_group_column, filename, sheet_name).
"""

import openpyxl
import os
import random
import string
import pandas as pd 
import csv

def create_empty_excel_file(filename):
    # Create a new Excel workbook
    workbook = openpyxl.Workbook()
    
    # Save the workbook to the specified filename
    workbook.save(filename)

    # Close the workbook
    workbook.close()

#TODO - results are problematic, gonna be fixed
#have to check input parameters as they are 3, but should be 2, no need filename with extension
def print_into_file(data_array, file_type, output_filename):
    if file_type not in ["xlsx", "csv"]:
        raise ValueError("Invalid file_type. Supported values are 'xlsx' or 'csv'.")

    if file_type == "xlsx":
        # Create a new Excel workbook
        workbook = openpyxl.Workbook()
        
        # Create a reference to the active sheet
        sheet = workbook.active

        # Print the data into the sheet
        for row_index, row_data in enumerate(data_array, start=1):
            for col_index, value in enumerate(row_data, start=1):
                sheet.cell(row=row_index, column=col_index, value=value)

        # Save the workbook as an Excel file
        workbook.save(output_filename)

    elif file_type == "csv":
        # Create a CSV file and print the data
        with open(output_filename, "w", newline="") as csv_file:
            csv_writer = csv.writer(csv_file)
            csv_writer.writerows(data_array)

def check_file_format(filename):
    
    if not os.path.exists(filename):
        raise FileNotFoundError(f"The file '{filename}' does not exist.")

    if filename.lower().endswith('.xlsx') or filename.lower().endswith('.csv'):
        # The file is in a supported format (Excel or CSV)
        return True
    else:
        # Unsupported file format
        print("Your data is not in a supported format. Please convert it to either XLSX or CSV.")
        exit(1)  # Terminate the program
def create_test_file(filename, num_columns, num_rows):
    # Create a new Excel workbook
    workbook = openpyxl.Workbook()

    # Create a reference to the active sheet
    sheet = workbook.active

    # Define headers for columns with alphabet letters
    headers = list(string.ascii_uppercase)

    for col_index in range(num_columns):

        # Fill the column with header
        for row_index in range(1, num_rows + 1):
            sheet.cell(row=row_index, column=col_index + 1, value=random.choice((headers)))

        # Fill the second column with a random year between 1950 and 2023
        if col_index == 1:
            for row_index in range(1, num_rows + 1):
                sheet.cell(row=row_index, column=col_index + 1, value=random.randint(1950, 2023))

        # Fill the third column with a random city name
        elif col_index == 2:
            cities = ["Ankara", "Istanbul", "Konya", "Izmir", "Atina", "Bursa"]
            for row_index in range(1, num_rows + 1):
                sheet.cell(row=row_index, column=col_index + 1, value=random.choice(cities))

        # Fill other columns with random numbers
        elif col_index > 2:
            for row_index in range(1, num_rows + 1):
                sheet.cell(row=row_index, column=col_index + 1, value=random.randint(1, 100))

    # Save the workbook to the specified filename
    workbook.save(filename)

    # Close the workbook
    workbook.close()

def random_sampler(filename, sampling_set_size, sheet_name=None):
    
    # Open the Excel file
    workbook = openpyxl.load_workbook(filename)
    
    # Access the sheet by name if provided, or use the first sheet by default
    if sheet_name:
        try:
            sheet = workbook[sheet_name]
        except KeyError:
            raise ValueError(f"Sheet '{sheet_name}' not found in the Excel file.")
    else:
        # Use the first sheet by default
        sheet = workbook.active
    
    # Get the total number of rows in the sheet
    num_rows = sheet.max_row
    
    if sampling_set_size >= num_rows:
        raise ValueError("Sampling set size cannot be greater than or equal to the number of rows.")
    
    # Generate a random sample of row indices
    sampled_rows_indices = random.sample(range(1, num_rows + 1), sampling_set_size)
    
    # Extract data from the sampled rows
    sampled_data = []
    #indices of selected rows. 
    rows_selected = []
    for row_index in sampled_rows_indices:
        row_data = [cell.value for cell in sheet[row_index]]
        sampled_data.append(row_data)
        rows_selected.append(row_index)
    return sampled_data

def stratified_sampler(filename, groupby_column_num, sample_size=None):
    # Check if the file exists
    if not os.path.exists(filename):
        raise FileNotFoundError(f"The file '{filename}' does not exist.")

    # Determine the file format (Excel or CSV)
    if filename.lower().endswith('.xlsx'):
        df = pd.read_excel(filename, header=None, skiprows=[0])
    elif filename.lower().endswith('.csv'):
        df = pd.read_csv(filename, header=None, skiprows=[0])
    else:
        raise ValueError("Unsupported file format. Please provide an Excel (.xlsx) or CSV (.csv) file.")

    # Get the specified column name based on the column number
    columns = df.columns
    if groupby_column_num < 1 or groupby_column_num > len(columns):
        raise ValueError("Invalid groupby_column_num. Column number is out of range.")
    groupby_column = columns[groupby_column_num - 1]
    
    # Set a default sample size
    if sample_size is None:
        sample_size = len(df) / 10

    # Group the data by the specified column
    grouped = df.groupby(groupby_column)

    sampled_data = []  # Array to store the sampled data
    percentages_per_group = []
    for group, group_data in grouped:
        group_size = len(group_data)
        percentage = (group_size / len(df)) * 100
        percentages_per_group.append(percentage)
        elements_to_select = int((percentage * sample_size) / 100)
        if elements_to_select > 0:
            # If elements_to_select is greater than the population size, sample with replacement
            if elements_to_select > group_size:
                sampled_group = group_data.sample(n=elements_to_select, replace=True)
            else:
                sampled_group = group_data.sample(n=elements_to_select, replace=False)
            sampled_data.append(sampled_group)

    # Combine the sampled data from all groups
    result = pd.concat(sampled_data).reset_index(drop=True)

    # Remove the indices from the printed result
    result_string = result.to_string(index=False)
    return result_string

def systematic_sampler(sampling_set_size, filename, sheet_name=None):
    # Open the Excel file
    workbook = openpyxl.load_workbook(filename)
    
    # Access the sheet by name if provided, or use the first sheet by default
    if sheet_name:
        try:
            sheet = workbook[sheet_name]
        except KeyError:
            raise ValueError(f"Sheet '{sheet_name}' not found in the Excel file.")
    else:
        # Use the first sheet by default
        sheet = workbook.active
    
    # Get the total number of rows in the sheet
    num_rows = sheet.max_row
    
    if num_rows == 0:
        raise ValueError("The sheet is empty. No data to sample.")

    # Calculate the step size to sample every nth item
    step_size = num_rows // sampling_set_size

    # Check if the sampling set size is valid
    if step_size < 1:
        raise ValueError("Sampling set size is too large for the given data.")

    # Initialize a list to store the sampled data
    sampled_data = []

    # Sample every nth item starting from the first row
    for row_index in range(1, num_rows + 1, step_size):
        row_data = [cell.value for cell in sheet[row_index]]
        sampled_data.append(row_data)

    return sampled_data

def cluster_sampler(sampling_set_size, sampling_group_column, filename, sheet_name=None):
    # Open the Excel file
    workbook = openpyxl.load_workbook(filename)
    
    # Access the sheet by name if provided, or use the first sheet by default
    if sheet_name:
        try:
            sheet = workbook[sheet_name]
        except KeyError:
            raise ValueError(f"Sheet '{sheet_name}' not found in the Excel file.")
    else:
        # Use the first sheet by default
        sheet = workbook.active
    
    # Get the total number of rows in the sheet
    num_rows = sheet.max_row
    
    if num_rows == 0:
        raise ValueError("The sheet is empty. No data to sample.")

    # Check if the provided group column exists in the sheet
    if sampling_group_column not in sheet[1]:
        raise ValueError(f"Group column '{sampling_group_column}' not found in the sheet.")

    # Create a list of unique groups based on the specified group column
    groups_array = list(set(sheet.cell(row=i, column=sheet[1].index(sampling_group_column) + 1).value for i in range(2, num_rows + 1)))

    # Select a random group from the groups_array
    selected_group = random.choice(groups_array)

    # Collect all rows with the selected group
    sampled_data = []
    for row_index in range(2, num_rows + 1):
        group_value = sheet.cell(row=row_index, column=sheet[1].index(sampling_group_column) + 1).value
        if group_value == selected_group:
            row_data = [cell.value for cell in sheet[row_index]]
            sampled_data.append(row_data)

    return sampled_data

#For future try-outs, if file uploaded is not in required file format, just terminate the program
# if not check_file_format("sample.xlsx"):  # Replace "sample.xlsx" with your filename
#     exit()  # Terminate the program
create_empty_excel_file("sample.xlsx")
create_test_file("sample.xlsx",3,250)
print(stratified_sampler("sample.xlsx", 3, 25))
print_into_file(stratified_sampler("sample.xlsx", 3, 25), "xlsx", "sonuc.xlsx")