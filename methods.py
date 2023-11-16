#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Thu Nov 16 22:51:36 2023

@author: pro
"""

import openpyxl
import os
import random
import string
import pandas as pd 

def create_test_file(filename, num_columns, num_rows):
    # Create a new Excel workbook
    workbook = openpyxl.Workbook()

    # Create a reference to the active sheet
    sheet = workbook.active

    # Define headers for columns with alphabet letters
    headers = list(string.ascii_uppercase)

    for col_index in range(num_columns):

        # Fill the column with header
        for row_index in range(1, num_rows + 1):
            sheet.cell(row=row_index, column=col_index + 1, value=random.choice((headers)))

        # Fill the second column with a random year between 1950 and 2023
        if col_index == 1:
            for row_index in range(1, num_rows + 1):
                sheet.cell(row=row_index, column=col_index + 1, value=random.randint(1950, 2023))

        # Fill the third column with a random city name
        elif col_index == 2:
            cities = ["Ankara", "Istanbul", "Konya", "Izmir", "Atina", "Bursa"]
            for row_index in range(1, num_rows + 1):
                sheet.cell(row=row_index, column=col_index + 1, value=random.choice(cities))

        # Fill other columns with random numbers
        elif col_index > 2:
            for row_index in range(1, num_rows + 1):
                sheet.cell(row=row_index, column=col_index + 1, value=random.randint(1, 100))

    # Save the workbook to the specified filename
    workbook.save(filename)

    # Close the workbook
    workbook.close()

def get_first_row_cells(filename):
    """
    Reads every cell value in the first row of an Excel file and returns them as a dictionary.

    Parameters:
        filename (str): The path to the Excel file.

    Returns:
        dict: A dictionary with column index as the key and cell value as the value.
    """
    # Check if the file exists
    if not os.path.exists(filename):
        return {"error": "[File not found]"}

    # Read the first row of the file
    if filename.lower().endswith('.xlsx'):
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.active
        first_row_cells = {col_index + 1: cell.value for col_index, cell in enumerate(sheet[1])}
        workbook.close()
    elif filename.lower().endswith('.csv'):
        with open(filename, 'r') as file:
            first_row_cells = {col_index + 1: value.strip() for col_index, value in enumerate(file.readline().strip().split(','))}
    else:
        return {"error": "[Unsupported file format]"}

    return first_row_cells

def random_sampler(filename, sampling_set_size, sheet_name=None):
    
    # Open the Excel file
    workbook = openpyxl.load_workbook(filename)
    
    # Access the sheet by name if provided, or use the first sheet by default
    if sheet_name:
        try:
            sheet = workbook[sheet_name]
        except KeyError:
            raise ValueError(f"Sheet '{sheet_name}' not found in the Excel file.")
    else:
        # Use the first sheet by default
        sheet = workbook.active
    
    # Get the total number of rows in the sheet
    num_rows = sheet.max_row
    
    if sampling_set_size >= num_rows:
        raise ValueError("Sampling set size cannot be greater than or equal to the number of rows.")
    
    # Generate a random sample of row indices
    sampled_rows_indices = random.sample(range(1, num_rows + 1), sampling_set_size)
    
    # Extract data from the sampled rows
    sampled_data = []
    #indices of selected rows. 
    rows_selected = []
    for row_index in sampled_rows_indices:
        row_data = [cell.value for cell in sheet[row_index]]
        sampled_data.append(row_data)
        rows_selected.append(row_index)
    return sampled_data

def stratified_sampler(filename, groupby_column_num, sample_size=None):
    # Check if the file exists
    if not os.path.exists(filename):
        raise FileNotFoundError(f"The file '{filename}' does not exist.")

    # Determine the file format (Excel or CSV)
    if filename.lower().endswith('.xlsx'):
        df = pd.read_excel(filename, header=None, skiprows=[0])
    elif filename.lower().endswith('.csv'):
        df = pd.read_csv(filename, header=None, skiprows=[0])
    else:
        raise ValueError("Unsupported file format. Please provide an Excel (.xlsx) or CSV (.csv) file.")

    # Get the specified column name based on the column number
    columns = df.columns
    if groupby_column_num < 1 or groupby_column_num > len(columns):
        raise ValueError("Invalid groupby_column_num. Column number is out of range.")
    groupby_column = columns[groupby_column_num - 1]
    
    # Set a default sample size
    if sample_size is None:
        sample_size = len(df) / 10

    # Group the data by the specified column
    grouped = df.groupby(groupby_column)

    sampled_data = []  # Array to store the sampled data
    percentages_per_group = []
    for group, group_data in grouped:
        group_size = len(group_data)
        percentage = (group_size / len(df)) * 100
        percentages_per_group.append(percentage)
        elements_to_select = int((percentage * sample_size) / 100)
        if elements_to_select > 0:
            # If elements_to_select is greater than the population size, sample with replacement
            if elements_to_select > group_size:
                sampled_group = group_data.sample(n=elements_to_select, replace=True)
            else:
                sampled_group = group_data.sample(n=elements_to_select, replace=False)
            sampled_data.append(sampled_group)

    # Combine the sampled data from all groups
    result = pd.concat(sampled_data).reset_index(drop=True)

    # Remove the indices from the printed result
    result_string = result.to_string(index=False)
    return result_string

def systematic_sampler(sampling_set_size, filename, sheet_name=None):
    # Open the Excel file
    workbook = openpyxl.load_workbook(filename)
    
    # Access the sheet by name if provided, or use the first sheet by default
    if sheet_name:
        try:
            sheet = workbook[sheet_name]
        except KeyError:
            raise ValueError(f"Sheet '{sheet_name}' not found in the Excel file.")
    else:
        # Use the first sheet by default
        sheet = workbook.active
    
    # Get the total number of rows in the sheet
    num_rows = sheet.max_row
    
    if num_rows == 0:
        raise ValueError("The sheet is empty. No data to sample.")

    # Calculate the step size to sample every nth item
    step_size = num_rows // sampling_set_size

    # Check if the sampling set size is valid
    if step_size < 1:
        raise ValueError("Sampling set size is too large for the given data.")

    # Initialize a list to store the sampled data
    sampled_data = []

    # Sample every nth item starting from the first row
    for row_index in range(1, num_rows + 1, step_size):
        row_data = [cell.value for cell in sheet[row_index]]
        sampled_data.append(row_data)

    return sampled_data

def cluster_sampler(sampling_set_size, sampling_group_column, filename, sheet_name=None):
    # Open the Excel file
    workbook = openpyxl.load_workbook(filename)
    
    # Access the sheet by name if provided, or use the first sheet by default
    if sheet_name:
        try:
            sheet = workbook[sheet_name]
        except KeyError:
            raise ValueError(f"Sheet '{sheet_name}' not found in the Excel file.")
    else:
        # Use the first sheet by default
        sheet = workbook.active
    
    # Get the total number of rows in the sheet
    num_rows = sheet.max_row
    
    if num_rows == 0:
        raise ValueError("The sheet is empty. No data to sample.")

    # Check if the provided group column exists in the sheet
    if sampling_group_column not in sheet[1]:
        raise ValueError(f"Group column '{sampling_group_column}' not found in the sheet.")

    # Create a list of unique groups based on the specified group column
    groups_array = list(set(sheet.cell(row=i, column=sheet[1].index(sampling_group_column) + 1).value for i in range(2, num_rows + 1)))

    # Select a random group from the groups_array
    selected_group = random.choice(groups_array)

    # Collect all rows with the selected group
    sampled_data = []
    for row_index in range(2, num_rows + 1):
        group_value = sheet.cell(row=row_index, column=sheet[1].index(sampling_group_column) + 1).value
        if group_value == selected_group:
            row_data = [cell.value for cell in sheet[row_index]]
            sampled_data.append(row_data)

    return sampled_data